import csv
import openpyxl
import argparse
import re
from datetime import datetime
from time import sleep
from tqdm import tqdm
from tkinter.filedialog import askopenfilename, asksaveasfilename
import argparse


def read_file(filepath):
    """Detects a file type and reads its contents accordingly"""
    if filepath.endswith('.csv'):
        return read_csv(filepath)
    elif filepath.endswith('.xlsx'):
        return read_excel(filepath)
    elif filepath.endswith('.log') or filepath.endswith('.txt'):
        return read_text_file(filepath)
    else:
        print("Unsupported file type")
        return []
def read_csv(filepath):
    """Reads a CSV file and return its contents as a list of lists"""
    with open(filepath, newline='') as csvfile:
        reader = csv.reader(csvfile)
        return [row for row in reader]
    

def read_excel(filepath):
    """Reads an Excel sheet and returns its contents as a list of lists"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    return data

def read_text_file(filepath):
    """Reads a text or log file line by line"""
    with open(filepath, 'r') as f:
        data = [line.strip().split(',')for line in f.readlines()]
        print(data)
        return (data)
    
def write_file(file_path, data):
    """Detects the file type and writes the content accordingly."""
    if file_path.endswith('.csv'):
        write_csv(file_path, data)
    elif file_path.endswith('.xlsx'):
        write_excel(file_path, data)
    elif file_path.endswith('.log') or file_path.endswith('.txt'):
        write_text_file(file_path, data)
    else:
        print("Unsupported file type.")

def write_csv(filepath, data):

    with open(filepath, mode='w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)
    print(f'Data successfully written to {filepath}')

def write_excel(filepath, data):
    print(data)
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(filepath)
    print(f'Data successfully written to {filepath}')


def write_text_file(filepath, data):
    with open(filepath, 'w') as file:
        for line in data:
            # Join lists into strings and write each line to the file
            if isinstance(line, list):
                file.write(''.join(str(item) for item in line) + '\n')
            else:
                file.write(str(line) + '\n')
    print(f'Data successfully written to {filepath}')


def remove_duplicates(rows):
    """Removes duplicate rows in data"""
    header, data_rows = rows[0], rows[1:]
    unique_rows = []
    seen = set()
    for row in tqdm(data_rows, desc="removing duplicates..."):
        row_tuple = tuple(row)
        if row_tuple not in seen:
            seen.add(row_tuple)
            unique_rows.append(row)
    return [header] + unique_rows


def convert_to_lowercase(rows):
    """Converts all data strings to lowercase"""
    header, data_rows = rows[0], rows[1:]
    new_rows = []
    for row in tqdm(data_rows, desc="Converting to lowercase..."):
        new_row = [cell.lower() if isinstance(cell, str) else cell for cell in row]
        new_rows.append(new_row)
    return [header] + new_rows


def remove_trailing_whitespace(rows):
    """Removes trailing whitespace from all data strings"""
    header, data_rows = rows[0], rows[1:]
    new_rows = []
    for row in tqdm(data_rows, desc="Trimming whitespace..."):
        new_row = [cell.strip() if isinstance(cell, str) else cell for cell in row]
        new_rows.append(new_row)
    return [header] + new_rows


def validate_email_format(rows, email_column):
    """Checks if an email is valid"""
    email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    
    new_rows = []
    header = rows[0]
    new_rows.append(header)
    
    # Check if an email header exists
    try:
        col_index = header.index(email_column)
        print()
    except ValueError:
        raise ValueError(f"Column '{email_column}' not found in the file header: {header}")
    
    for row in tqdm(rows[1:], desc="Validating emails..."):
        if re.match(email_regex, row[col_index]):
            new_rows.append(row)
        else:
            row[col_index] = "Invalid email"
            new_rows.append(row)
        
    return new_rows


def standardize_phone_numbers(rows, phone_column):
    """Standardizes phone numbers per North American
    format: (xxx) xxx-xxxx"""
    phone_regex = r'(\d{3})[`^\d]*(\d{3})[`^\d]*(\d{4})'
    
    new_rows = []
    header = rows[0]
    new_rows.append(header)
    col_index = header.index(phone_column)

    for row in tqdm(rows[1:], desc="Standardizing phone numbers..."):
        if re.search(phone_regex, row[col_index]):
            row[col_index] = re.sub(phone_regex, r'(\1) \2-\3', row[col_index])
        else:
            row[col_index] = "Invalid phone number"
        new_rows.append(row)
    return new_rows


def validate_and_format_dates(rows, date_format="%Y-%m-%d"):
    """Converts dates from `MM/DD/YYYY` to `YYYY-MM-DD`."""
    new_rows = []

    for row in rows:
        new_row = []

        for cell in row:
            if isinstance(cell, str):
                try:
                    parsed_date = datetime.strptime(cell, "%m/%d/%Y")
                    new_row.append(parsed_date.strftime(date_format))
                except ValueError:
                    new_row.append(cell)
            else:
                new_row.append(cell)
        new_rows.append(new_row)
    return new_rows


def check_missing_values(data, column_name):
    """Checks for missing values in a specified column
    and flags them"""
    header, rows = data[0], data[1:]

    try:
        col_index = header.index(column_name)
    except ValueError:
        raise ValueError(f'Column {column_name} not found in header file {header}')
    
    for row in tqdm(rows, desc=f"Checking for missing values in {column_name}"):
        if row[col_index] == None or row[col_index] == '':
            row[col_index] = "Missing value"
    return [header] + rows


def validate_data_type(data, column_name, expected_type):
    """Validates data types in the specified column."""
    header, rows = data[0], data[1:]
    try:
        col_index = header.index(column_name)
    except ValueError:
        raise ValueError(f"Column '{column_name}' not found in the file header: {header}")

    # Define the type validation functions
    type_checks = {
        'int': lambda val: isinstance(val, int),
        'float': lambda val: isinstance(val, float),
        'string': lambda val: isinstance(val, str),
        'date': lambda val: isinstance(val, datetime)
    }

    # Check if the expected_type is valid
    if expected_type not in type_checks:
        raise ValueError(
            f"Unsupported type: '{expected_type}'. Supported types are: {list(type_checks.keys())}")

    # Validate data type for each row
    for row in rows:
        value = row[col_index]
        if not type_checks[expected_type](value):
            row[col_index] = "Invalid type"
    
    return [header] + rows

def parse_arguments():
    parser = argparse.ArgumentParser(description="Text-Based Data Cleaning Assistant")
    parser.add_argument('--trim_whitespace', action='store_true', help='Trim leading and trailing whitespace')
    parser.add_argument('--lowercase', action='store_true', help='Convert text to lowercase')
    parser.add_argument('--remove_duplicates', action='store_true', help='Remove duplicate rows')
    parser.add_argument('--validate_email', help='Check email is valid')
    parser.add_argument('--standardize_phone', help='Standardize phone numbers')
    parser.add_argument('--format_date', action='store_true', help='Validate and format dates')
    parser.add_argument("--check_missing", help="Check for missing values in a specific column", type=str)
    parser.add_argument("--validate_type", help="Validate data type in a specific column", nargs=2, metavar=('column_name', 'type'))
    return parser.parse_args()


if __name__ == "__main__":
    filepath = askopenfilename()
    data = read_file(filepath)

    args = parse_arguments()

    if data:
        print("Cleaning data... ")
        if args.trim_whitespace:
            data = remove_trailing_whitespace(data)
        if args.lowercase:
            data = convert_to_lowercase(data)
        if args.remove_duplicates:
            data = remove_duplicates(data)
        if args.validate_email:
            data = validate_email_format(data, args.validate_email)
        if args.standardize_phone:
            data = standardize_phone_numbers(data, args.standardize_phone)
        if args.format_date:
            data = validate_and_format_dates(data)
        if args.check_missing:
            data = check_missing_values(data, args.check_missing)
        if args.validate_type:
            column_name, expected_type = args.validate_type
            print(f"Validating data type '{expected_type}' in column '{column_name}'...")
            data = validate_data_type(data, column_name, expected_type)
    else:
        print("No data to process!")

    sleep(5)
        
    filepath = asksaveasfilename()
    write_file(filepath, data)